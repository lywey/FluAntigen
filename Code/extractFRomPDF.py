import re
import openpyxl
from datetime import datetime
# 输入数据
# 输入数据99/99/2999 2999-99-99 5a B,H3N2,H1N1
savefile_path = r"C:\Users\刘青博\Downloads\Who2024-9 表H3-23 H3N2.xlsx"
value_column_13 = "2024_9"
value_column_12 = "Table 23. (Additional) Antigenic analyses of influenza A(H3N2) viruses (Guinea Pig RBC with 20nm Oseltamivir) 2024-09-10"
n=12#即滴度的列数
input_data = """
REFERENCE VIRUSES
A/Croatia/10136/RV/2023  2a.3a.1(J.2) 2023-12-04 SIAT3 320 320 320 160 160 160 160 320 320 320 40 40
A/Croatia/10136/RV/2023  2a.3a.1(J.2) 2999-99-99 E3(Am1Al2) 640 640 640 640 320 640 640 640 640 640 80 80
A/Netherlands/10563/2023  2a.3a.1(J.2) 2023-11-10 K-MIX2/SIAT3 160 320 320 320 160 160 320 640 320 320 40 80
A/Lisboa/216/2023  2a.3a.1(J.2) 2023-12-15 E3(Am1Al2) 320 320 320 320 160 320 640 640 640 640 40 80
A/Slovenia/49/2024  2a.3a.1(J.2) 2024-01-08 MDCKx/SIAT3 160 40 80 160 1280 80 80 80 80 80 <40 640
A/District Of Columbia/27/2023  2a.3a.1(J.2) 2999-99-99 SIAT2 320 320 320 320 160 320 320 640 640 640 80 40
A/District Of Columbia/27/2023  2a.3a.1(J.2) 2999-99-99 Spf2E2 320 320 320 640 320 320 320 320 640 640 80 40
A/Idaho/69/2023  2a.3a.1(J.2) 2999-99-99 SPFE1 160 320 160 40 <40 80 160 320 80 160 80 <40
A/Oklahoma/05/2024  2a.3a.1(J.2) 2999-99-99 SIAT1 160 40 80 320 2560 80 160 80 80 160 <40 1280
A/Colorado/06/2024  2a.3a.1(J.2) 2999-99-99 SIAT1 160 160 320 320 160 160 320 640 640 640 80 80
A/Nevada/32/2023  2a.3a.1(J.2) 2999-99-99 SIAT2 640 320 640 640 640 640 640 1280 640 1280 80 160
A/Shanghai-Fengxian/1912/2023  2a.3a.1(J.2) 2999-99-99 E2/E10 320 640 640 640 320 640 1280 1280 1280 1280 160 160
A/Sydney/878/2023  2a.3a.1(J.2) 2999-99-99 SIAT1 640 640 640 640 320 320 640 1280 640 1280 80 160
A/Victoria/2997/2023  2a.3a.1(J.2) 2999-99-99 SIAT2 160 320 320 160 160 160 320 640 320 320 80 40
A/Canberra/302/2023  2a.3a.1(J.2) 2999-99-99 E2 640 640 640 640 320 640 640 1280 1280 1280 80 320
A/Canberra/309/2023  2a.3a.1(J.2) 2999-99-99 E2 160 320 320 640 160 320 320 640 640 320 40 80
A/Victoria/2776/2023  2a.3a.1(J.2) 2999-99-99 E2 640 640 640 1280 320 320 640 640 1280 1280 40 320
IVR-252 A/Canberra/302/2023  2a.3a.1(J.2) 2999-99-99 E2/D10 160 160 160 320 160 320 640 320 320 1280 40 160
IVR-253 A/Switzerland/8649/2023  2a.3a.1(J.2) 2999-99-99 E3/D7 160 160 320 320 160 160 320 640 320 320 40 40
IVR-254 A/Canberra/373/2023  2a.3a.1(J.2) 2999-99-99 E4/D7 80 160 160 160 160 160 320 320 320 320 <40 <40
IVR-258 A/Victoria/2997/2023  2a.3a.1(J.2) 2999-99-99 E4/D6 320 640 640 640 160 320 320 320 640 320 <40 80
A/Stockholm/SE24-53070/2024  2a.3a.1(J.2) 2024-04-22 SIAT1/SIAT1 160 320 320 320 160 160 320 640 320 320 80 40
A/BosniaAndHerzegovina/502/2024  2a.3a.1(J.2) 2024-01-30 SIAT1 160 160 320 160 80 80 160 320 160 160 40 <40
A/Catalonia/NSVH102371258/2024  2a.3a.1(J.2) 2024-06-25 SIAT2/SIAT1 160 320 320 320 160 320 320 640 320 320 40 80
A/Catalonia/NSVH102337351/2024  2a.3a.1(J.2) 2024-05-07 SIAT2/SIAT1 160 320 320 320 160 160 320 320 320 640 80 40
A/Idaho/69/2023  2a.3a.1(J.2) 2999-99-99 S2 ND ND ND ND ND ND ND ND ND ND ND ND
A/Oklahoma/05/2024  2a.3a.1(J.2) 2999-99-99 Spf2E1 ND ND ND ND ND ND ND ND ND ND ND ND

"""
input_data2 = """
A/Croatia/ A/Croatia/ A/Netherlands A/Lisboa/ A/Slovenia A/DistrictOf A/DistrictOf A/Colorado A/Nevada A/Sydney A/Idaho/ A/Oklahoma
10136/RV/2023 0136/RV/2023 /10563/2023 216/2023 /49/2024 umbia/27/2023 mbia/27/2023 /06/2024 /32/2023 /878/2023 69/2023 /05/2024
SIAT Egg SIAT Egg SIAT SIAT Egg SIAT SIAT SIAT SIAT EGG
F06/24 F16/24 F08/24 F15/24 F11/24 F24-022 F24-037 F24-035 F24-012 A9825 F24-033 F24-044
2a.3a.1(J.2) 2a.3a.1(J.2) 2a.3a.1(J.2) 2a.3a.1(J.2) 2a.3a.1(J.2) 2a.3a.1(J.2) 2a.3a.1(J.2) 2a.3a.1(J.2) 2a.3a.1(J.2) 2a.3a.1(J.2) 2a.3a.1(J.2) 2a.3a.1(J.2)
"""
# 创建一个新的工作簿
workbook_2 = openpyxl.Workbook()
sheet_2 = workbook_2.active
sheet_2.title = "Virus Data"
# 定义列标题
headers = [
    "virusIsoalte", "virusStrain", "virusStrain genetic group", "CollectionDate",
    "virusstrain passage history", "serumIsolate", "serumStrain passage",
    "serum ferret number", "serumstrain genetic group", "serumstrain",
    "HI titre", "source", "who report", "other information"
]
# 将列标题写入第一行
for col_num, header in enumerate(headers, 1):
    sheet_2.cell(row=1, column=col_num, value=header)
# 定义保存文件的路径，使用原始字符串
   ## 请替换成导出的数据集文件路径
# 保存 Excel 文件
workbook_2.save(savefile_path)
print(f"Excel 文件已保存到: {savefile_path}")        #生成一个空的导出数据集表格

# 当前行索引
row_index = 2

# 遍历每一行，提取最后 13 个数据作为滴度数据
for line in input_data.strip().splitlines():
    # 如果行包含 'TEST VIRUSES' 则输出空行
    if 'TEST VIRUSES' in line:
        # sheet_2.cell(row=row_index, column=11, value="")  # 写入空行
        # row_index += 1
        print("")  # 输出空行
        continue  # 跳过后续处理

    # 使用正则表达式找到数字、ND、<或>符号跟随的数字
    match = re.findall(r'(ND|[<>]?\d+)', line)

    # 只提取最后 13 个滴度数据
    if len(match) >= n:
        titer_data = match[-n:]  # 取最后 13 个
        print("\t".join(titer_data))
        # 将滴度数据用 Tab 分隔符连接并输出
        for titer in titer_data:
            sheet_2.cell(row=row_index, column=11, value=titer)
            row_index += 1


# 保存 Excel 文件
workbook_2.save(savefile_path)
print(f"Excel 文件已保存到: {savefile_path}")


row_index = 2

# 改进后的正则表达式：匹配病毒名称，直到遇到 "数字+小写字母" 的组合时停止
pattern = r'^([A-Za-z0-9/_\-\(\). ]+)(?=\s(?:\d+[a-z]|no\s+seq|pending))'
# 标记当前的病毒类型（REFERENCE VIRUSES 或 TEST VIRUSES）
current_type = None
# 保存输出结果的列表
results = []
# 遍历输入数据
for line in input_data.strip().splitlines():
    line = line.strip()

    # 检查是否是 REFERENCE VIRUSES 或 TEST VIRUSES
    if "REFERENCE VIRUSES" in line or "TEST VIRUSES" in line:
        current_type = line  # 直接记录类型
        results.append(current_type)
        continue

    # 匹配病毒名称部分
    match = re.search(pattern, line)
    if match:
        virus_name = match.group(1)  # 只捕获病毒名称部分
        # 将提取出的病毒名称处理并添加到列表
        results.append(virus_name)
        for _ in range(n):
            sheet_2.cell(row=row_index, column=1, value=virus_name)
            row_index += 1
    else:
        print(f"No match found in line: {line}")

# 检查如果结果列表中只有 REFERENCE VIRUSES 和 TEST VIRUSES，说明病毒名称没有被正确捕获
if len(results) <= 2:
    print("Error: No virus names were captured. Please check the input data and the regular expression.")

# 输出处理后的病毒名称
for result in results:
    print(result)

# 保存 Excel 文件
workbook_2.save(savefile_path)
print(f"Excel 文件已保存到: {savefile_path}")



row_index = 2
# 定义正则表达式提取包含 "一个数字 + 小写字母" 或 "no seq" 的字符串
pattern = r'\b\S*\d+[a-z]\S*\b|\bno\s+seq\b|\bno\s+sequence\b|\bpending\b'
# pattern = r'\b\S*\d+[a-z][A-Z]*\S*\b|\bno\s+seq\b|\bpending\b'
# pattern = r'\b\S*\d+[a-z][A-Z]*\)?\S*\b|\bno\s+seq\b|\bpending\b'
# 保存匹配结果的列表
matches = []
# 遍历输入数据，查找匹配项
# for line in input_data.strip().splitlines():
#     if 'TEST VIRUSES' in line:
#         matches.append('')  # 遇到 TEST VIRUSES 行输出空行
#     else:
#         line_matches = re.findall(pattern, line)
#         if line_matches:
#             matches.extend(line_matches)
#             for match in line_matches:
#                 for _ in range(n):
#                     sheet_2.cell(row=row_index, column=3, value=match)  # 写入第三列
#                     row_index += 1
# 遍历输入数据，查找匹配项
for line in input_data.strip().splitlines():
    if 'TEST VIRUSES' in line:
        matches.append('')  # 遇到 TEST VIRUSES 行输出空行
    else:
        line_matches = re.findall(pattern, line)
        if line_matches:
            matches.extend(line_matches)
            for match in line_matches:
                # 检查是否包含左括号 `(` 且不包含右括号 `)`
                if '(' in match and ')' not in match:
                    match += ')'  # 在末尾添加右括号 `)`

                # 写入第三列
                for _ in range(n):
                    sheet_2.cell(row=row_index, column=3, value=match)
                    row_index += 1
# 输出匹配的结果
for match in matches:
    print(match)

# 保存 Excel 文件
workbook_2.save(savefile_path)
print(f"Excel 文件已保存到: {savefile_path}")



row_index = 2
# 定义正则表达式提取包含日期格式的字符串
pattern = r'\b(\d{4}-\d{2}-\d{2}|unknown)\b'

# 保存匹配结果的列表
matches = []

# 遍历输入数据，查找匹配项
for line in input_data.strip().splitlines():
    # 如果行包含 'TEST VIRUSES' 则输出空行
    if 'TEST VIRUSES' in line:
        matches.append('')
    else:
        line_matches = re.findall(pattern, line)
        if line_matches:
            matches.extend(line_matches)
            for match in line_matches:
                for _ in range(n):
                    sheet_2.cell(row=row_index, column=4, value=match)
                    row_index += 1

# 输出匹配的结果
for match in matches:
    print(match)
# 定义保存文件的路径

# 保存 Excel 文件
workbook_2.save(savefile_path)
print(f"Excel 文件已保存到: {savefile_path}")
#
#



row_index = 2
# 定义正则表达式匹配日期
date_pattern = r'\b(\d{4}-\d{2}-\d{2}|unknown)\b'
# 定义正则表达式匹配滴度数据（包括 "<" 开头）
titer_pattern = r'\b(?:ND|<\d+|\d+)\b'
# 定义正则表达式匹配标识符，包括可选的数字后缀（如 10-3）
identifier_pattern = r'[A-Za-z0-9/]+(?:\s+\d+-\d+)?'

# 保存结果的列表
results = []

# 遍历每一行，提取所需数据
for line in input_data.strip().splitlines():
    # 如果行包含 'TEST VIRUSES' 则输出空行
    if 'TEST VIRUSES' in line:
        # results.append('')
        continue
    else:
        # 查找日期
        date_match = re.search(date_pattern, line)
        if date_match:
            date_end = date_match.end()  # 获取日期匹配的结束位置
            # 查找日期之后的标识符部分
            identifier_match = re.search(identifier_pattern, line[date_end:])
            if identifier_match:
                result = identifier_match.group().strip()
                # 检查是否后面紧跟 "<" 或其他符号
                remaining_text = line[identifier_match.end() + date_end:].strip()
                if not remaining_text.startswith("<") and not remaining_text.startswith("ND"):
                    results.append(result)
                    for _ in range(n):
                        sheet_2.cell(row=row_index, column=5, value=result)
                        row_index += 1
                else:
                    results.append(result.split()[0])
                    for _ in range(n):
                        sheet_2.cell(row=row_index, column=5,
                                     value=result.split()[0])
                        row_index += 1

# 输出结果
for result in results:
    print(result)
# 定义保存文件的路径


# 保存 Excel 文件
workbook_2.save(savefile_path)
print(f"Excel 文件已保存到: {savefile_path}")


# 初始化计数器
reference_viruses_count = 0
test_viruses_count = 0
current_section = None

# 遍历输入数据，统计每部分的行数
for line in input_data.strip().splitlines():
    if 'REFERENCE VIRUSES' in line:
        current_section = 'REFERENCE'
    elif 'TEST VIRUSES' in line:
        current_section = 'TEST'
    elif current_section == 'REFERENCE':
        reference_viruses_count += 1
    elif current_section == 'TEST':
        test_viruses_count += 1

# 输出结果
print(f"REFERENCE VIRUSES 下的毒株行数: {reference_viruses_count}")
print(f"TEST VIRUSES 下的毒株行数: {test_viruses_count}")
print(f"excel边界行数: {7+reference_viruses_count+test_viruses_count}")



# 输入数据


# 将输入的多行数据分割成列表
lines = input_data2.strip().splitlines()

# 第一部分：处理病毒名称和编号拼接
first_line_list = lines[0].split()  # 第一行数据
second_line_list = lines[1].split()  # 第二行数据
third_line_list = lines[2].split()  # 第三行数据
fourth_line_list = lines[3].split()  # 第四行数据
fifth_line_list = lines[4].split()  # 第五行数据
# 设定行索引从第二行开始（第一行是标题行）
row_index = 2
n = reference_viruses_count+test_viruses_count  # 可以根据需要调整

# 检查行数是否匹配
if len(first_line_list) == len(second_line_list):
    for _ in range(n):  # 循环n次
        for i in range(len(first_line_list)):
            combined_string_6 = first_line_list[i] + "/" + second_line_list[i]
            combined_string_7 = third_line_list[i]

            # 过滤掉不需要的 "Egg" 条目

            sheet_2.cell(row=row_index, column=6, value=combined_string_6)
            sheet_2.cell(row=row_index, column=7, value=combined_string_7)
            row_index += 1

else:
    print("Error: The lines have different numbers of elements.")
row_index = 2
n = reference_viruses_count+test_viruses_count

for _ in range(n):  # 循环n次
    for i in range(len(fourth_line_list)):
        combined_string_8 = fourth_line_list[i]
        combined_string_9 = fifth_line_list[i]
        sheet_2.cell(row=row_index, column=8, value=combined_string_8)
        sheet_2.cell(row=row_index, column=9, value=combined_string_9)
        row_index += 1
for line in lines[3:]:
    formatted_line = '\t'.join(line.split())
    print(formatted_line)
# 保存文件路径

# 保存 Excel 文件
workbook_2.save(savefile_path)
print(f"Excel 文件已保存到: {savefile_path}")
current_row = 2  # 从第2行开始（跳过表头）
for row_idx in range(current_row, sheet_2.max_row + 1):
    sheet_2.cell(row=row_idx, column=13, value=value_column_13)
workbook_2.save(savefile_path)
current_row = 2  # 从第2行开始（跳过表头）
for row_idx in range(current_row, sheet_2.max_row + 1):
    sheet_2.cell(row=row_idx, column=12, value=value_column_12)
workbook_2.save(savefile_path)