import re
import openpyxl
from datetime import datetime
# 99/99/2999 2999-99-99 5a B,H3N2,H1N1，REFERENCE VIRUSES  TEST VIRUSES
value_column_14 = "2019_12"
value_column_12 = "Table 3-1. Antigenic analysis of A(H1N1)pdm09 viruses by HI"
parts = value_column_12.split('Table ')
if len(parts) > 1:
    second_part = parts[1].split('.')
    if len(second_part) > 1:
        table_number = second_part[0]
        virus_part = second_part[1].split('A(')
        if len(virus_part) > 1:
            virus_info = virus_part[1].split(')')
            virus_type = virus_info[0] if virus_info else None
            # virus_type = "H3N2"
savefile_path = fr"D:\CDC\ EXCEl\{value_column_14} table{table_number} {virus_type}.xlsx"   #ND ND ND

input_data1 = """			
REFERENCE VIRUSES
 A/Michigan/45/2015 6B.1 2015-09-07 E3/E4 2560 640 640 2560 2560 2560 1280 2560 2560 1280
 A/Bayern/69/2009  2009-07-01 MDCK5/MDCK1 80 320 320 40 160 160 80 320 40 80
 A/Lviv/N6/2009  2009-10-27 MDCK4/SIAT1/MDCK3 320 640 640 160 640 320 160 640 160 320
 A/Slovenia/2903/2015  6B.1 2015-10-26 E4/E2 2560 640 1280 5120 5120 5120 1280 5120 1280 2560
 A/Paris/1447/2017 6B.1A 2017-10-20 MDCK1/MDCK3 640 320 160 1280 2560 2560 640 2560 1280 1280
 A/Switzerland/2656/2017 6B.1A 2017-12-21 E5/E3 2560 640 640 2560 2560 5120 2560 5120 2560 2560
 A/Switzerland/3330/2017  6B.1A5B 2017-12-20 E6/E2 640 320 160 1280 1280 1280 1280 640 640 640
 A/Norway/3433/2018 6B.1A5A 2018-10-30 MDCK3 320 80 40 320 640 640 320 640 320 320
 A/Ireland/84630/2018 6B.1A6 2018-11-28 MDCK1/MDCK3 640 160 160 1280 1280 1280 640 640 1280 640
 A/Brisbane/02/2018 6B.1A1 2018-01-04 E3/E1 1280 320 320 1280 2560 2560 1280 1280 1280 1280
 TEST VIRUSES
 A/England/292/2019 6B.1A5A 2019-09-06 SIAT1/MDCK1 320 160 80 320 640 320 320 1280 320 320
 A/England/296/2019 6B.1A5A 2019-09-13 SIAT1/MDCK1 640 320 160 640 1280 640 640 1280 640 640
 A/England/298/2019 6B.1A5A 2019-10-10 SIAT1/MDCK1 640 160 80 640 1280 640 320 1280 640 320
 A/Norway/2316/2019 6B.1A5A 2019-10-16 MDCK1 640 160 160 640 1280 640 640 1280 640 320
 A/Norway/2347/2019 6B.1A5A 2019-10-23 MDCK1 320 80 40 160 320 160 160 640 160 160
 A/Norway/2368/2019 6B.1A5A 2019-11-03 SIAT1/MDCK1 640 160 160 640 640 640 640 1280 640 640
 A/Norway/2317/2019 6B.1A5B 2019-10-21 MDCK1 640 80 40 640 640 320 320 640 640 320
 A/Norway/2412/2019 6B.1A5B 2019-11-06 MDCK1 1280 320 160 1280 1280 1280 640 2560 1280 640
 A/Norway/2406/2019 6B.1A5B 2019-11-08 MDCK1 640 160 80 640 640 640 640 1280 640 640
 A/England/297/2019 6B.1A6 2019-10-08 SIAT1/MDCK1 640 320 160 640 1280 640 640 1280 640 640
 A/Norway/2471/2019 6B.1A7 2019-11-14 MDCK1 640 320 160 1280 2560 1280 1280 2560 1280 1280
 A/Norway/2475/2019 6B.1A7 2019-11-16 MDCK1 1280 320 160 1280 1280 1280 640 2560 1280 1280
"""
input_data2 = """     
A/Mich A/Bayern A/Lviv A/Slov A/Paris A/Swit A/Swit A/Norway A/Ire A/Bris
45/15 69/09 N6/09 2903/2015 1447/17 2656/17 3330/17 3433/18 84630/18 02/18
Egg MDCK MDCK Egg MDCK Egg Egg MDCK MDCK Egg
F31/16*1 F09/15*1 F13/18*1 NIBF48/16*1 F03/18*2 F20/18*1 F23/18*1 F04/19*1 F08/19*1 F09/19*1
6B.1 FFF FFF 6B.1 6B.1A 6B.1A 6B.1A5B 6B.1A5A 6B.1A6 6B.1A1
"""
# Split the input multi-line data into lists   V1A V1A V1A V1A V1A V1A V1A V1A FFF FFF FFF
# FFF FFF FFF FFF FFF FFF FFF FFF FFF FFF
# Use splitlines to split the data into lines
lines = input_data2.strip().splitlines()
# Select the first line (you can also select other lines, because all lines should have the same number of columns)
first_line = lines[0]
# Split the first line by space to calculate the number of columns
columns = first_line.split()
n = len(columns)  # Get the number of columns

print("Number of columns:", n)

date_pattern = r'\b(\d{4}-\d{2}-\d{2}|unknown)\b'
# date_pattern = r'\b(\d{2}/\d{2}/\d{4}|unknown)\b'
# Process the data in rows
lines = input_data1.strip().split('\n')
# Output a list to store the processed results
output_lines = []
# Traverse each line
for line in lines:
    # Match date position
    date_match = re.search(date_pattern, line)
    if date_match:
        date_position = date_match.start()
        # Get the content before the date and split it by space
        before_date = line[:date_position].strip().split()
        # Determine whether the column before the date is a single number or a combination of numbers and letters
        column_before_date = before_date[-1]
        # if re.match(r'^\d+[A-Za-z]?$|^\d$|^\d+[A-Za-z]?\.[\w.∆-]+$', column_before_date):
        if re.match(r'^\d{1}[A-Za-z]{2}$|^\d+[A-Za-z]?$|^\d$|^\d+[A-Za-z]?\.[\w.-]+$|1A\(∆2\)$|1A\(∆3\)$|1A\(∆3\)B$|1A\(∆2\)B$|\bV1[A-Za-z0-9.]*\b|pending', column_before_date):
            #If yes, slice and remove the last element, the last element is "number" or "number plus letter combination"
            virus_name = " ".join(before_date[:-1])  # virus name
            suffix = f"9x{column_before_date}"  # concatenate 9x and the last
        else:
            # If not, the virus name contains all the contents
            virus_name = " ".join(before_date)
            suffix = "9x"  # 只添加 5a,改为xx
        # Construct the new line to avoid repeating the strain name
        rest_of_line = line[date_position:].strip()
        new_line = f"{virus_name} {suffix} {rest_of_line}".strip()
        output_lines.append(new_line)
    else:
        # If no date is found, keep the original line
        output_lines.append(line)
# Output the processed result
output_data = '\n'.join(output_lines)
input_data = output_data
# input_data = """
#
# """
# Create a new workbook
workbook_2 = openpyxl.Workbook()
sheet_2 = workbook_2.active
sheet_2.title = "Virus Data"
# Define column headers
headers = [
    "VirusIsoalte", "VirusStrain", "VirusStrain genetic group", "CollectionDate",
    "VirusStrain Passage History", "SerumIsolate", "SerumStrain Passage History",
    "Ferret Number", "SerumStrain Genetic Group", "SerumStrain",
    "HI titre", "SourceName", "SourceType", "PublishedDate"
]
# Write column headers to the first row
for col_num, header in enumerate(headers, 1):
    sheet_2.cell(row=1, column=col_num, value=header)
# Save the Excel file
workbook_2.save(savefile_path)
print(f"Excel file saved to: {savefile_path}")        #Generate an empty export data set table
# Current row index
row_index = 2
# n = 10  # Number of titer columns
# Traverse each line to extract the last 13 data points as titer data
for line in input_data.strip().splitlines():
    # If the line contains 'TEST VIRUSES', output an empty line
    if 'TEST VIRUSES' in line:
        # sheet_2.cell(row=row_index, column=11, value="")  # Write an empty line
        # row_index += 1
        print("")  # Output empty line
        continue  # Skip further processing

    # Use regular expression to find numbers, ND, < or > symbols followed by numbers
    # match = re.findall(r'(ND|[<>]?\d+)', line)
    match = re.findall(r'(ND|[<>≤]?\d+|[<>])', line)

    # Only extract the last 13 titer data points
    if len(match) >= n:
        titer_data = match[-n:]  # Take the last n
        print("\t".join(titer_data))
        # Join titer data with Tab separator and output
        for titer in titer_data:
            sheet_2.cell(row=row_index, column=11, value=titer)
            row_index += 1

# Save the Excel file
workbook_2.save(savefile_path)
print(f"Excel file saved to: {savefile_path}")


row_index = 2

# Improved regular expression: Match the virus name until encountering a combination of "number + lowercase letter"
# pattern = r'^([A-Za-z0-9/_\-\(\). ]+)(?=\s(?:\d+[a-z]|no\s+seq|pending))'
pattern = r'^([A-Za-z0-9/_\-\(\).+‑ ₀-₉⁰-⁹]+)(?=\s(?:\d+[a-z]|no\s+seq|pending|\bV1[A-Za-z0-9.]*\b))'
# pattern = r'\b5a\b'
# Mark the current virus type (REFERENCE VIRUSES or TEST VIRUSES)
current_type = None
# List to store output results
results = []
# Traverse the input data
for line in input_data.strip().splitlines():
    line = line.strip()

    # Check if it is REFERENCE VIRUSES or TEST VIRUSES
    if "REFERENCE VIRUSES" in line or "TEST VIRUSES" in line:
        current_type = line  # Directly record the type
        results.append(current_type)
        continue

    # Match the virus name part
    match = re.search(pattern, line)
    if match:
        virus_name = match.group(1)  # Capture only the virus name part
        # Process and add the captured virus name to the list
        results.append(virus_name)
        for _ in range(n):
            sheet_2.cell(row=row_index, column=1, value=virus_name)
            row_index += 1
    else:
        print(f"No match found in line: {line}")

# Check if the results list only contains REFERENCE VIRUSES and TEST VIRUSES, indicating no virus names were captured
if len(results) <= 2:
    print("Error: No virus names were captured. Please check the input data and the regular expression.")

# Output the processed virus names
for result in results:
    print(result)

# Save the Excel file
workbook_2.save(savefile_path)
print(f"Excel file saved to: {savefile_path}")



row_index = 2
# Define regular expression to extract strings containing "a number + lowercase letter" or "no seq"
# pattern = r'\b\S*\d+[a-z]\S*\b|\bno\s+seq\b|\bpending\b'
# pattern = r'\b\S*\d+[a]\S*\b|\bno\s+seq\b|\bpending\b'
# pattern = r'\b\S*\d+[x]\S*\b|\bno\s+seq\b|\bpending\b'
pattern = r'\b\S*\d+[x]\S*\b|\bno\s+seq\b|\bpending\b|1A\(∆3\)\b|1A\(∆3\)B\b|1A\(∆2\)\b|\bV1[A-Za-z0-9.]*\b'
# pattern = r'\b5a\b'
# List to store the matching results
matches = []
# Iterate through the input data to find matches
for line in input_data.strip().splitlines():
    if 'TEST VIRUSES' in line:
        matches.append('')  # Output an empty line when encountering 'TEST VIRUSES' line
    else:
        line_matches = re.findall(pattern, line)
        if line_matches:
            matches.extend(line_matches)
            for match in line_matches:
                for _ in range(n):
                    sheet_2.cell(row=row_index, column=3, value=match)  # Write to the third column
                    row_index += 1

# Output the matched results
for match in matches:
    print(match)

# Save the Excel file
workbook_2.save(savefile_path)
print(f"Excel file has been saved to: {savefile_path}")



row_index = 2
# Define regular expression to extract strings containing date format
pattern = r'\b(\d{4}-\d{2}-\d{2}|unknown)\b'
# pattern = r'\b(\d{2}/\d{2}/\d{4}|unknown)\b'
# List to store the matching results
matches = []

# Iterate through the input data to find matches
for line in input_data.strip().splitlines():
    # Output an empty line if the line contains 'TEST VIRUSES'
    if 'TEST VIRUSES' in line:
        matches.append('')
    else:
        line_matches = re.findall(pattern, line)
        if line_matches:
            matches.extend(line_matches)
            for match in line_matches:
                for _ in range(n):
                    sheet_2.cell(row=row_index, column=4, value=match)
                    row_index += 1

# Output the matched results
for match in matches:
    print(match)
# Define the path to save the file

# Save the Excel file
workbook_2.save(savefile_path)
print(f"Excel file has been saved to: {savefile_path}")
#
#

row_index = 2
# Define regular expression to match dates
date_pattern = r'\b(\d{4}-\d{2}-\d{2}|unknown)\b'
# date_pattern = r'\b(\d{2}/\d{2}/\d{4}|unknown)\b'
# Define regular expression to match titer data (including those starting with "<")
titer_pattern = r'\b(?:ND|<\d+|\d+)\b'
# Define regular expression to match identifiers, including optional number suffixes (10-3)
# identifier_pattern = r'[A-Za-z0-9/]+(?:\s+\d+-\d+)?'
identifier_pattern = r'[A-Za-z0-9/-]+(?:\s+\d+-\d+)?'
# List to store the results
results = []

# Iterate through each line to extract the required data
for line in input_data.strip().splitlines():
    # If the line contains 'TEST VIRUSES', skip it
    if 'TEST VIRUSES' in line:
        # results.append('')
        continue
    else:
        # Find the date
        date_match = re.search(date_pattern, line)
        if date_match:
            date_end = date_match.end()  # Get the end position of the date match
            # Find the identifier part after the date
            identifier_match = re.search(identifier_pattern, line[date_end:])
            if identifier_match:
                result = identifier_match.group().strip()
                # # Check if there is any "<" or other symbols following it
                remaining_text = line[identifier_match.end() + date_end:].strip()
                if not remaining_text.startswith("<") and not remaining_text.startswith("ND"):
                    results.append(result)
                    for _ in range(n):
                        sheet_2.cell(row=row_index, column=5, value=result)
                        row_index += 1
                else:
                    results.append(result.split()[0])
                    for _ in range(n):
                        sheet_2.cell(row=row_index, column=5,
                                     value=result.split()[0])
                        row_index += 1

# Output the results
for result in results:
    print(result)
# Define the path to save the file


# Save the Excel file
workbook_2.save(savefile_path)
print(f"Excel file saved to: {savefile_path}")


# Initialize counters
reference_viruses_count = 0
test_viruses_count = 0
current_section = None

# Iterate through the input data to count the lines in each section
for line in input_data.strip().splitlines():
    if 'REFERENCE VIRUSES' in line:
        current_section = 'REFERENCE'
    elif 'TEST VIRUSES' in line:
        current_section = 'TEST'
    elif current_section == 'REFERENCE':
        reference_viruses_count += 1
    elif current_section == 'TEST':
        test_viruses_count += 1

# Output the results
print(f"Number of strains under REFERENCE VIRUSES: {reference_viruses_count}")
print(f"Number of strains under TEST VIRUSES: {test_viruses_count}")
print(f"Excel boundary row count: {7+reference_viruses_count+test_viruses_count}")


#Egg Egg Egg MDCK Egg MDCK MDCK MDCK MDCK MDCK FFFF
# Split the input data into lines   V1A V1A V1A V1A V1A V1A V1A V1A
lines = input_data2.strip().splitlines()
#V1A V1A V1A
# Part one: Process virus names and numbers
first_line_list = lines[0].split()  # First line data
second_line_list = lines[1].split()  # Second line data
third_line_list = lines[2].split()  # Third line data
fourth_line_list = lines[3].split()  # Fourth line data
fifth_line_list = lines[4].split()  # Fifth line data
# fifth_line_list = re.split(r'\s{4,}', lines[4])
# fifth_line = lines[4]
# fixed_column_width = 15
# fifth_line_list = re.split(r'(?<=\S)\s{4,}|(?<=\s)\s{4,}(?=\S)', lines[4])
# fifth_line_list = []
# for i in range(0, len(fifth_line), fixed_column_width):
#     column = fifth_line[i:i + fixed_column_width].strip()
#     fifth_line_list.append(column)
# print("Total columns:", len(fifth_line_list))
# Set row index starting from the second row (the first row is the header row)
row_index = 2
n = reference_viruses_count+test_viruses_count  # Can be adjusted as needed
# Check if the row counts match
if len(first_line_list) == len(second_line_list):
    for _ in range(n):  # loop n times
        for i in range(len(first_line_list)):
            combined_string_6 = first_line_list[i] + "/" + second_line_list[i]
            combined_string_7 = third_line_list[i]
            # Filter out unwanted "Egg" entries
            sheet_2.cell(row=row_index, column=6, value=combined_string_6)
            sheet_2.cell(row=row_index, column=7, value=combined_string_7)
            row_index += 1

else:
    print("Error: The lines have different numbers of elements.")
row_index = 2
n = reference_viruses_count+test_viruses_count

for _ in range(n):  # Repeat n times
    for i in range(len(fourth_line_list)):
        combined_string_8 = fourth_line_list[i]
        # combined_string_9 = fifth_line_list[i]
        sheet_2.cell(row=row_index, column=8, value=combined_string_8)
        # sheet_2.cell(row=row_index, column=9, value=combined_string_9)
        row_index += 1
# Iterate through the fourth and fifth line data
row_index = 2
for i in range(n):
    for i in range(len(fifth_line_list)):
        combined_string_9 = fifth_line_list[i]
        sheet_2.cell(row=row_index, column=9, value=combined_string_9)
        row_index += 1
for line in lines[3:]:
    formatted_line = '\t'.join(line.split())
    print(formatted_line)
# Save file path
# Save the Excel file
workbook_2.save(savefile_path)
print(f"Excel file saved to: {savefile_path}")
current_row = 2  # Start from row 2 (skip the table header)
for row_idx in range(current_row, sheet_2.max_row + 1):
    sheet_2.cell(row=row_idx, column=14, value=value_column_14)
workbook_2.save(savefile_path)
current_row = 2  # Start from row 2 (skip the table header)
for row_idx in range(current_row, sheet_2.max_row + 1):
    sheet_2.cell(row=row_idx, column=12, value=value_column_12)
workbook_2.save(savefile_path)